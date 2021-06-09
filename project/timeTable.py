from datetime import date, datetime, timedelta
from pathlib import Path

import openpyxl
import requests
import xlrd
from bs4 import BeautifulSoup
from dateutil import parser

import dbFunctions
import parameters


def init_timetable():
    if needToUpdate():
        get_timetable()


def get_timetable():
    # Адрес сайта расписания РТУ МИРЭА.
    url = 'https://www.mirea.ru/schedule/'

    # Получаем страницу, парсим через BeautifulSoup.
    page = requests.get(url)
    soup = BeautifulSoup(page.text, "html.parser")

    # Получаем куски тегов с ссылками.
    result = soup.find("div", {"class": "rasspisanie"}). \
        find(string="Институт информационных технологий"). \
        find_parent("div"). \
        find_parent("div"). \
        findAll('a', class_='uk-link-toggle')

    # Словарь.
    dictTimetables = {1: None, 2: None, 3: None}

    # Выводим ссылки на расписание.
    print("Links to timetables:")
    for x in result:
        print("  link: " + str(x["href"]))
        if "ИИТ_1к" in x["href"] and "весна" in x["href"]:
            print("    Downloading started.")
            f = open("file1.xlsx", "wb")
            resp = requests.get(x["href"])
            f.write(resp.content)
            dictTimetables[1] = f
            print("    Downloading completed.")
        elif "ИИТ_2к" in x["href"] and "весна" in x["href"]:
            print("    Downloading started.")
            f = open("file2.xlsx", "wb")
            resp = requests.get(x["href"])
            f.write(resp.content)
            dictTimetables[2] = f
            print("    Downloading completed.")
        elif "ИИТ_3к" in x["href"] and "весна" in x["href"]:
            print("    Downloading started.")
            f = open("file3.xlsx", "wb")
            resp = requests.get(x["href"])
            f.write(resp.content)
            dictTimetables[3] = f
            print("    Downloading completed.")
        else:
            print("    Ignored.")
    insert_timetable_into_db(dictTimetables)


# Вставка расписания в базу данных, скачивание файлов.
def insert_timetable_into_db(dictTimetables):
    print("Inserting to database started.")

    sqlite_connection = dbFunctions.sqlite_connection
    cursor = sqlite_connection.cursor()

    command = "DELETE FROM timetable_pairs;"
    cursor.execute(command)

    xlsx = openpyxl.load_workbook('file1.xlsx')
    ws = xlsx.active
    for row in ws.iter_rows():
        for cell in row:
            if str(cell.value).find("ИАБО-") != -1 \
                    or str(cell.value).find("ИВБО-") != -1 \
                    or str(cell.value).find("ИКБО-") != -1 \
                    or str(cell.value).find("ИМБО-") != -1 \
                    or str(cell.value).find("ИНБО-") != -1:
                print("  found group: " + str(cell.value))
                record_timetable_into_db(str(cell.value), int(cell.row), int(cell.column), ws)
    xlsx = openpyxl.load_workbook('file2.xlsx')
    ws = xlsx.active
    for row in ws.iter_rows():
        for cell in row:
            if str(cell.value).find("ИАБО-") != -1 \
                    or str(cell.value).find("ИВБО-") != -1 \
                    or str(cell.value).find("ИКБО-") != -1 \
                    or str(cell.value).find("ИМБО-") != -1 \
                    or str(cell.value).find("ИНБО-") != -1:
                print("  found group: " + str(cell.value))
                record_timetable_into_db(str(cell.value), int(cell.row), int(cell.column), ws)
    xlsx = openpyxl.load_workbook('file3.xlsx')
    ws = xlsx.active
    for row in ws.iter_rows():
        for cell in row:
            if str(cell.value).find("ИАБО-") != -1 \
                    or str(cell.value).find("ИВБО-") != -1 \
                    or str(cell.value).find("ИКБО-") != -1 \
                    or str(cell.value).find("ИМБО-") != -1 \
                    or str(cell.value).find("ИНБО-") != -1:
                print("  found group: " + str(cell.value))
                record_timetable_into_db(str(cell.value), int(cell.row), int(cell.column), ws)

    command = "INSERT INTO timetable_updates (session) VALUES (\'" + date.today().strftime('%Y-%m-%d') + "\');"
    cursor.execute(command)
    cursor.close()
    sqlite_connection.commit()

    print(parameters.OKGREEN + "Database updated" + parameters.ENDC)


# Запись расписания в базу данных.
def record_timetable_into_db(group_name, cell_row, cell_col, ws):
    print("    Recording started.")

    sqlite_connection = dbFunctions.sqlite_connection
    cursor = sqlite_connection.cursor()

    for days in range(1, 7):
        for pair in range(1, 7):
            for k in range(1, 3):
                even = False
                if k == 1:
                    even = False
                else:
                    even = True
                pair_name = str(ws.cell(((cell_row + 2) + 2 * (pair - 1) + (k - 1)) + (days - 1) * 12, cell_col).value)
                pair_type = str(
                    ws.cell(((cell_row + 2) + 2 * (pair - 1) + (k - 1)) + (days - 1) * 12, cell_col + 1).value)
                pair_teacher = str(
                    ws.cell(((cell_row + 2) + 2 * (pair - 1) + (k - 1)) + (days - 1) * 12, cell_col + 2).value)
                pair_room = str(
                    ws.cell(((cell_row + 2) + 2 * (pair - 1) + (k - 1)) + (days - 1) * 12, cell_col + 3).value)
                command = "INSERT INTO timetable_pairs VALUES(\'" + group_name + "\', \'" + str(days) + "\', \'" + str(
                    pair) + \
                          "\', \'" + str(even) + "\', \'" + pair_name + "\', \'" + pair_type + "\', \'" + \
                          pair_teacher + "\', \'" + pair_room + "\');"
                cursor.execute(command)

    cursor.close()
    sqlite_connection.commit()
    print("    Recording completed.")


# Проверка на обновление расписания.
def needToUpdate():
    sqlite_connection = dbFunctions.sqlite_connection
    cursor = sqlite_connection.cursor()

    try:
        command = "SELECT MAX(id) FROM timetable_updates;"
        c = cursor.execute(command)
        for row in c:
            if row[0] is None:
                return True
    except:
        return True

    command = "SELECT * FROM timetable_updates ORDER BY id DESC LIMIT 1;"
    c = cursor.execute(command)
    for row in c:
        if int(row[0]) < 1:
            return True
        else:
            date = parser.parse(row[1])
            if date.date() < date.today().date():
                return True
            else:
                print(parameters.OKGREEN + "Database was recently updated (" + str(
                    date.today().date()) + ")" + parameters.ENDC)
                return False
    return False

# Ф-ция нахождения группы.
def findGroup(group_name):
    print("Trying to find group " + group_name)
    sqlite_connection = dbFunctions.sqlite_connection
    cursor = sqlite_connection.cursor()
    command = "SELECT DISTINCT group_name FROM timetable_pairs WHERE group_name = \'" + group_name + "\';"
    c = cursor.execute(command)
    for row in c:
        if row[0] is None:
            return False
        else:
            return True
    return False

# Получить расписание группы.
def get_group_timetable(group_name, period):
    global val

    if period == "today":
        if datetime.today().weekday() + 1 == 7:
            return "🎉 Отдыхаем, выходной!"
        if (datetime.today().isocalendar()[1] - 5) % 2 == 0:
            pair_even = True
        else:
            pair_even = False
        val = "⏰ Расписание на сегодня:\n"
        for i in range(1, 7):
            pair_number = i
            day_number = datetime.today().weekday() + 1
            val = val + str(get_pair(group_name, pair_even, pair_number, day_number))
    elif period == "tomorrow":
        if (datetime.today() + timedelta(days=1)).weekday() + 1 == 7:
            return "🎉 Отдыхаем, выходной!"
        if ((datetime.today() + timedelta(days=1)).isocalendar()[1] - 5) % 2 == 0:
            pair_even = True
        else:
            pair_even = False
        val = "⏰ Расписание на завтра:\n"
        for i in range(1, 7):
            pair_number = i
            day_number = (datetime.today() + timedelta(days=1)).weekday() + 1
            val = val + str(get_pair(group_name, pair_even, pair_number, day_number))
    elif period == "nextweek":
        print((datetime.today()).isocalendar()[1])
        print((datetime.today() + timedelta(days=7)).isocalendar()[1])
        if ((datetime.today() + timedelta(days=7)).isocalendar()[1] - 5) % 2 == 0:
            pair_even = True
        else:
            pair_even = False
        val = "⏰ Расписание на следующую неделю:"
        for day_val in range(0, 7):
            val = val + "\n👉 " + get_day_rus(day_val) + ":\n"
            for i in range(1, 7):
                pair_number = i
                day_number = day_val + 1
                if day_number == 7:
                    val = val + "🎉 Выходной.\n"
                    break
                else:
                    val = val + str(get_pair(group_name, pair_even, pair_number, day_number))
    elif period == "currentweek":
        if ((datetime.today()).isocalendar()[1] - 5) % 2 == 0:
            pair_even = True
        else:
            pair_even = False
        val = "⏰ Расписание на текущую неделю:"
        for day_val in range(0, 7):
            val = val + "\n👉 " + get_day_rus(day_val) + ":\n"
            for i in range(1, 7):
                pair_number = i
                day_number = day_val + 1
                if day_number == 7:
                    val = val + "🎉 Выходной.\n"
                    break
                else:
                    val = val + str(get_pair(group_name, pair_even, pair_number, day_number))
    elif period in ("понедельник", "вторник", "среда", "четверг", "пятница", "суббота", "воскресенье"):
        val = "⏰ Расписание - " + period + ":\n"
        for i in range(1, 7):
            for k in range(1, 3):
                if k == 2:
                    pair_even = True
                    val = val + "⏩"
                else:
                    pair_even = False
                    val = val + "▶"
                pair_number = i
                day_number = get_day_number(period) + 1
                if day_number == 7:
                    val = val[:-1] + "🎉 Выходной.\n"
                    return val
                else:
                    val = val + str(get_pair(group_name, pair_even, pair_number, day_number))
        val = val + "\n\n▶ - нечетная неделя\n⏩ - четная неделя "

    return val

# Форматировать расписание.
def format_timetable(pair_number, pair_name=None, pair_type=None, pair_teacher=None, pair_room=None):
    val = " "
    if pair_number == 1:
        val = val + "1️⃣"
    elif pair_number == 2:
        val = val + "2️⃣"
    elif pair_number == 3:
        val = val + "3️⃣"
    elif pair_number == 4:
        val = val + "4️⃣"
    elif pair_number == 5:
        val = val + "5️⃣"
    elif pair_number == 6:
        val = val + "6️⃣"

    if pair_name is None:
        val = val + " -\n"
    else:
        if pair_room == "None":
            pair_room = "-"
        if pair_type == "None":
            pair_type = "-"
        if pair_teacher == "None":
            pair_teacher = "-"
        val = val + " " + pair_name.replace('\n', ' ').replace('\r', '') + ", " + pair_type.replace('\n', ' ').replace(
            '\r', '') + ", " + \
              pair_teacher.replace('\n', ' ').replace('\r', '') + ", " + pair_room.replace('\n', ' ').replace('\r',
                                                                                                              '') + "." + "\n"
    return val

# Получить пару.
def get_pair(group_name, pair_even, pair_number, day_number):
    sqlite_connection = dbFunctions.sqlite_connection
    cursor = sqlite_connection.cursor()
    command = "SELECT pair_name, pair_type, pair_teacher, pair_room FROM timetable_pairs WHERE pair_even = \'" + str(
        pair_even) + "\' AND pair_number = " + \
              str(pair_number) + " AND day_number = " + str(day_number) + " AND group_name = \'" + str(
        group_name) + "\';"
    print(command)
    c = cursor.execute(command)

    for row in c:
        if row[0] == "None":
            return format_timetable(pair_number)
        pair_name = str(row[0])
        pair_type = str(row[1])
        pair_teacher = str(row[2])
        pair_room = str(row[3])
        return format_timetable(pair_number, pair_name, pair_type, pair_teacher, pair_room)

# Перевод с английского на русский дня недели.
def get_day_rus(day_week):
    if day_week == 0:
        return "Понедельник"
    elif day_week == 1:
        return "Вторник"
    elif day_week == 2:
        return "Среда"
    elif day_week == 3:
        return "Четверг"
    elif day_week == 4:
        return "Пятница"
    elif day_week == 5:
        return "Суббота"
    elif day_week == 6:
        return "Воскресенье"

# Получить номер дня по названию.
def get_day_number(day_week):
    if day_week == "понедельник":
        return 0
    elif day_week == "вторник":
        return 1
    elif day_week == "среда":
        return 2
    elif day_week == "четверг":
        return 3
    elif day_week == "пятница":
        return 4
    elif day_week == "суббота":
        return 5
    elif day_week == "воскресенье":
        return 6

# Получить список учителей.
def get_teachers(teacher_name):
    sqlite_connection = dbFunctions.sqlite_connection
    cursor = sqlite_connection.cursor()
    command = "SELECT DISTINCT pair_teacher FROM timetable_pairs WHERE pair_teacher LIKE \'%" + str(
        teacher_name).capitalize() + "%\';"
    c = cursor.execute(command)
    lst = []
    for row in c:
        print(row[0])
        words = row[0].split()
        for item in list(map(' '.join, zip(words[:-1], words[1:]))):
            lst.append(item)
    res = []
    for i in lst:
        if i not in res:
            res.append(i)
    for i in range(0, len(res)):
        if res[i].find(str(teacher_name).capitalize()) == -1:
            res[i] = "None"
            continue
        words = res[i].split()
        finalstr = ""
        if words[0].find(".") == -1:
            finalstr = words[0] + " " + words[1]
        else:
            finalstr = words[1] + " " + words[0]
        res[i] = finalstr

    lst.clear()
    for i in res:
        if i not in lst and i != "None":
            lst.append(i)
    return lst

# Вывести расписание преподавателя на сегодня.
def get_teacher_timetable_today(teacher_name):
    sqlite_connection = dbFunctions.sqlite_connection
    cursor = sqlite_connection.cursor()
    if ((datetime.today() + timedelta(days=0)).isocalendar()[1] - 5) % 2 == 0:
        pair_even = True
    else:
        pair_even = False
    day_number = datetime.today().weekday() + 1
    command = "SELECT DISTINCT pair_number, pair_name, pair_type, pair_room FROM timetable_pairs WHERE pair_teacher " \
              "LIKE \'%" + str(
        teacher_name).title() + "%\' AND pair_even = \'" + str(pair_even) + "\' AND day_number = " + str(
        day_number) + " ORDER BY pair_number;"
    print(command)
    c = cursor.execute(command)
    return_str = "🔔 Расписание преподавателя " + str(teacher_name).title() + " на сегодня:\n"
    wasInCycle = False
    for row in c:
        return_str = return_str + format_timetable(row[0], row[1], row[2], "-", row[3])
        wasInCycle = True
    if not wasInCycle:
        return_str = "☺ Сегодня у преподавателя нет пар!"
    return return_str

# Вывести расписание преподавателя на завтра.
def get_teacher_timetable_tomorrow(teacher_name):
    sqlite_connection = dbFunctions.sqlite_connection
    cursor = sqlite_connection.cursor()
    if ((datetime.today() + timedelta(days=1)).isocalendar()[1] - 5) % 2 == 0:
        pair_even = True
    else:
        pair_even = False
    day_number = datetime.today().weekday() + 2
    command = "SELECT DISTINCT pair_number, pair_name, pair_type, pair_room FROM timetable_pairs WHERE pair_teacher " \
              "LIKE \'%" + str(
        teacher_name).title() + "%\' AND pair_even = \'" + str(pair_even) + "\' AND day_number = " + str(
        day_number) + " ORDER BY pair_number;"
    print(command)
    c = cursor.execute(command)
    return_str = "🔔 Расписание преподавателя " + str(teacher_name).title() + " на завтра:\n"
    wasInCycle = False
    for row in c:
        return_str = return_str + format_timetable(row[0], row[1], row[2], "-", row[3])
        wasInCycle = True
    if not wasInCycle:
        return_str = "☺ Завтра у преподавателя нет пар!"
    return return_str

# Вывести расписание преподавателя на эту неделю.
def get_teacher_timetable_this_week(teacher_name):
    sqlite_connection = dbFunctions.sqlite_connection
    cursor = sqlite_connection.cursor()
    if ((datetime.today() + timedelta(days=0)).isocalendar()[1] - 5) % 2 == 0:
        pair_even = True
    else:
        pair_even = False
    return_str = "🔔 Расписание преподавателя " + str(teacher_name).title() + " на эту неделю:"
    for day_number in range(0, 7):
        command = "SELECT DISTINCT pair_number, pair_name, pair_type, pair_room FROM timetable_pairs WHERE " \
                  "pair_teacher LIKE \'%" + str(
            teacher_name).title() + "%\' AND pair_even = \'" + str(pair_even) + "\' AND day_number = " + str(
            day_number + 1) + " ORDER BY pair_number;"
        print(command)
        c = cursor.execute(command)
        return_str = return_str + "\n👉 " + get_day_rus(day_number) + "\n"
        wasInCycle = False
        for row in c:
            return_str = return_str + format_timetable(row[0], row[1], row[2], "-", row[3])
            wasInCycle = True
        if not wasInCycle:
            return_str = return_str + "☺ В этот день у преподавателя нет пар!\n"

    return return_str

# Вывести расписание преподавателя на следующую неделю.
def get_teacher_timetable_next_week(teacher_name):
    sqlite_connection = dbFunctions.sqlite_connection
    cursor = sqlite_connection.cursor()
    if ((datetime.today() + timedelta(days=7)).isocalendar()[1] - 5) % 2 == 0:
        pair_even = True
    else:
        pair_even = False
    return_str = "🔔 Расписание преподавателя " + str(teacher_name).title() + " на следующую неделю:"
    for day_number in range(0, 7):
        command = "SELECT DISTINCT pair_number, pair_name, pair_type, pair_room FROM timetable_pairs WHERE " \
                  "pair_teacher LIKE \'%" + str(
            teacher_name).title() + "%\' AND pair_even = \'" + str(pair_even) + "\' AND day_number = " + str(
            day_number + 1) + " ORDER BY pair_number;"
        print(command)
        c = cursor.execute(command)
        return_str = return_str + "\n👉 " + get_day_rus(day_number) + "\n"
        wasInCycle = False
        for row in c:
            return_str = return_str + format_timetable(row[0], row[1], row[2], "-", row[3])
            wasInCycle = True
        if not wasInCycle:
            return_str = return_str + "☺ В этот день у преподавателя нет пар!\n"

    return return_str
